"""Parser for the user's Finance.xlsx historical data.

The workbook has monthly sheets (Jan, Feb, Mar, APRIL, MAY, JUN, JULY,
August, Sep, Oct, Nov, Dec) with inconsistent casing.  Each sheet has:

- SALARY section: income rows (col 1 = name, col 2 = amount)
- EXPENSES section: fixed bill rows (col 0 = "OK" status, col 1 = name, col 2 = amount)
- Variable spending grid: category headers in cols 6-14, individual amounts below
- Category headers: Food, Restaurants, Clothes/cosmetic, Household,
  Fun things, Sport/wellness, Travelling, Bullshit, Important

This parser extracts income + fixed expenses + variable spending as
transaction dicts ready for ImportService.execute_import.
"""

import re
from datetime import date

import openpyxl

# Month sheet names mapped to month numbers (1-12)
MONTH_NAMES = {
    "jan": 1, "feb": 2, "mar": 3, "april": 4,
    "may": 5, "jun": 6, "july": 7, "august": 8,
    "sep": 9, "oct": 10, "nov": 11, "dec": 12,
}

# Variable spending category columns (0-indexed from col 6)
VARIABLE_CATEGORIES = [
    "Food", "Restaurants", "Clothes/cosmetic", "Household",
    "Fun things", "Sport/wellness", "Travelling", "Bullshit", "Important",
]


def parse_finance_excel(file_path: str, year: int) -> list[dict]:
    """Parse Finance.xlsx and return flat list of transaction dicts.

    Each dict has: date, description, amount, type, section, category_name.
    """
    wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
    all_transactions = []

    for sheet_name in wb.sheetnames:
        month_num = _match_month(sheet_name)
        if month_num is None:
            continue

        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        txns = _parse_monthly_sheet(rows, year, month_num)
        all_transactions.extend(txns)

    wb.close()
    return all_transactions


def _match_month(sheet_name: str) -> int | None:
    """Map a sheet name to a month number (1-12) or None."""
    lower = sheet_name.strip().lower()
    return MONTH_NAMES.get(lower)


def _parse_monthly_sheet(rows: list[tuple], year: int, month: int) -> list[dict]:
    """Extract transactions from a single monthly sheet."""
    transactions = []
    section = None  # 'salary', 'expenses', or 'variable'
    variable_cols = None  # column indices for variable spending grid

    for row_idx, row in enumerate(rows):
        # Ensure row has enough elements to check
        if len(row) < 3:
            continue

        col0 = _cell_str(row[0])
        col1 = _cell_str(row[1])
        col2 = row[2] if len(row) > 2 else None

        # Detect section markers
        if col1 == "SALARY":
            section = "salary"
            continue
        if col1 == "EXPENCES" or col1 == "EXPENSES":
            section = "expenses"
            continue

        # Detect variable spending header row (contains category names)
        if _is_variable_header(row):
            section = "variable"
            variable_cols = _find_variable_columns(row)
            continue

        # Detect end of useful data
        if col1 and ("SAVINGS" in col1.upper() or "TOTAL" in col1.upper()):
            if section == "expenses":
                section = None
                continue

        # Parse salary/income rows
        if section == "salary" and col1 and col2 is not None:
            amount = _to_float(col2)
            if amount and amount > 0:
                transactions.append({
                    "date": _make_date(year, month, 1),
                    "description": col1.strip(),
                    "amount": amount,
                    "type": "income",
                    "section": "income",
                    "category_name": "Salary" if "salary" in col1.lower() or "zuzi" in col1.lower() or "maks" in col1.lower() else col1.strip(),
                })

        # Parse fixed expense rows
        elif section == "expenses" and col1 and col2 is not None:
            amount = _to_float(col2)
            if amount and amount > 0:
                # Skip summary rows
                if col1.startswith("Costs ") or col1 == "TOTAL":
                    continue
                transactions.append({
                    "date": _make_date(year, month, 1),
                    "description": col1.strip(),
                    "amount": amount,
                    "type": "expense",
                    "section": "fixed",
                    "category_name": col1.strip(),
                })

        # Parse variable spending rows
        elif section == "variable" and variable_cols:
            for cat_name, col_idx in variable_cols.items():
                if col_idx < len(row):
                    val = row[col_idx]
                    amount = _to_float(val)
                    if amount and amount > 0:
                        transactions.append({
                            "date": _make_date(year, month, 15),
                            "description": cat_name,
                            "amount": amount,
                            "type": "expense",
                            "section": "variable",
                            "category_name": cat_name,
                        })

        # If we hit the totals row for variable (large values in all variable cols), stop
        if section == "variable" and variable_cols and _is_totals_row(row, variable_cols):
            section = None

    return transactions


def _is_variable_header(row: tuple) -> bool:
    """Check if a row contains the variable spending category headers."""
    text = " ".join(str(c or "") for c in row).lower()
    return "food" in text and "restaurants" in text and "household" in text


def _find_variable_columns(row: tuple) -> dict[str, int]:
    """Find column indices for each variable spending category."""
    result = {}
    for i, cell in enumerate(row):
        cell_str = _cell_str(cell)
        if not cell_str:
            continue
        for cat in VARIABLE_CATEGORIES:
            if cat.lower() in cell_str.lower():
                result[cat] = i
                break
    return result


def _is_totals_row(row: tuple, variable_cols: dict[str, int]) -> bool:
    """Check if this is the totals row (all variable cols have large numbers)."""
    non_zero = 0
    for col_idx in variable_cols.values():
        if col_idx < len(row):
            val = _to_float(row[col_idx])
            if val and val > 500:
                non_zero += 1
    return non_zero >= 5


def _cell_str(cell) -> str:
    """Convert cell to stripped string, empty string if None."""
    if cell is None:
        return ""
    return str(cell).strip()


def _to_float(value) -> float:
    """Convert value to float, return 0.0 on failure."""
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    text = re.sub(r"[A-Za-z$€£¥Kč]", "", text)
    text = text.replace(" ", "").replace(",", ".")
    try:
        return abs(float(text))
    except ValueError:
        return 0.0


def _make_date(year: int, month: int, day: int) -> date:
    """Create a date, clamping day to valid range."""
    import calendar
    max_day = calendar.monthrange(year, month)[1]
    return date(year, month, min(day, max_day))
